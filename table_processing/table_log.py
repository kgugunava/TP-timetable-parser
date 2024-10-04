import openpyxl


def ras(cst_i, i):
    prom_arr = []
    for col in worksheet.iter_cols(4, worksheet.max_column):
        for row in range(75, 8, -1):
            data = col[row].value
            if data != None:
                p_data = ""
                # print(data, end= "\n**********\n")
                for iter in data:
                    p_data += iter
                prom_arr += p_data
                if "=" in p_data or f"КНТ {i}" in p_data:
                    cst_i += [prom_arr]
                    prom_arr = []

    for iter_2 in range(6):
        k = "".join(cst_i[iter_2])
        cst_i[iter_2] = k.split(". ")[::-1]

    return cst_i

wookbook = openpyxl.load_workbook("test.xlsx")
worksheet = wookbook.active

def data_set():
    cst_1 = []
    cst_2 = []
    cst_3 = []
    cst_4 = []
    cst_5 = []
    cst_6 = []
    cst_7 = []
    cst_8 = []
    cst_9 = []

    cst_1 = ras(cst_1, 1)[:6]
    cst_2 = ras(cst_2, 2)[:6]
    cst_3 = ras(cst_3, 3)[:6]
    cst_4 = ras(cst_4, 4)[:6]
    cst_5 = ras(cst_5, 5)[:6]
    cst_6 = ras(cst_6, 6)[:6]
    cst_7 = ras(cst_7, 7)[:6]
    cst_8 = ras(cst_8, 8)[:6]
    cst_9 = ras(cst_9, 9)[:6]

    rasp = [cst_1, cst_2, cst_3, cst_4, cst_5, cst_6, cst_7, cst_8, cst_9]

    return rasp


for i in data_set():
    for j in i:
        print(j)
    print("*****")


